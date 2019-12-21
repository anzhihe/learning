import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print(sheet.max_row, sheet.max_column)

for rowOfCellObjs in sheet['A1':'C3']:
    for cellObj in rowOfCellObjs:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')