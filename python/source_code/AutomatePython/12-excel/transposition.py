import openpyxl
import sys
from openpyxl.utils.cell import get_column_letter

if len(sys.argv) < 2:
    print('usage: py transposition.py [xlsx]')
    exit()

xlsxFile = sys.argv[1]

# Open xlsx and read it
wbRead = openpyxl.load_workbook(xlsxFile)
sheetRead = wbRead.get_active_sheet()

# Create sheet to be write
wbWrite = openpyxl.Workbook()
sheetWrite = wbWrite.get_active_sheet()

# Transposition
for i in range(1, sheetRead.max_row+1):
    for j in range(1, sheetRead.max_column+1):
        coordRead = get_column_letter(j)+str(i)
        coordWrite = get_column_letter(i)+str(j)
        sheetWrite[coordWrite] = sheetRead[coordRead].value

# save sheet to file
wbWrite.save(xlsxFile+'.trans.xlsx')


