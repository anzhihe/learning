import openpyxl
import sys
from openpyxl.utils.cell import get_column_letter

# arguments validation
textList = []
excelName = ''

# create workbook
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

# open each file and save to workbook
for i in range(1, sheet.max_column+1):
    with open(str(i)+'.txt', 'w') as f:
    for j in range(1, sheet.max_row+1):
        coord = get_column_letter(fileSeq) + str(rowSeq)
        f.write(sheet[coord].value)
    f.close()
