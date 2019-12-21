import openpyxl
import sys
from openpyxl.utils.cell import get_column_letter

if len(sys.argv) < 4:
    print('usage: py blankRowInserter.py [afterTheRow] [blankRowsNumber] [xlsx]')
    exit()

afterTheRow = int(sys.argv[1])
blankRowsNumber = int(sys.argv[2])
xlsxFile = sys.argv[3]

# Open xlsx and read it
wbRead = openpyxl.load_workbook(xlsxFile)
sheetRead = wbRead.get_active_sheet()

# Create sheet to be write
wbWrite = openpyxl.Workbook()
sheetWrite = wbWrite.get_active_sheet()

# copy first afterTheRow rows
for i in range(1, afterTheRow+1):
    for j in range(1, sheetRead.max_column+1):
        coord = get_column_letter(j)+str(i)
        sheetWrite[coord] = sheetRead[coord].value
# copy latter rows

for i in range(afterTheRow+1, sheetRead.max_row+1):
    for j in range(1, sheetRead.max_column+1):
        coordWrite = get_column_letter(j)+str(i+blankRowsNumber)
        coordRead = get_column_letter(j)+str(i)
        sheetWrite[coordWrite] = sheetRead[coordRead].value

# save sheet to file
wbWrite.save(xlsxFile+'.blank.xlsx')


