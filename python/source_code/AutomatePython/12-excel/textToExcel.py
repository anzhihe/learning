import openpyxl
import sys
from openpyxl.utils.cell import get_column_letter

# arguments validation
textList = []
excelName = ''

# create workbook
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

# open each file and save to workbook
fileSeq = 0
for file in textList:
    fileSeq += 1
    with open(file) as f:
        rowSeq = 1
        for line in f.readlines():
           coord = get_column_letter(fileSeq) + str(rowSeq)
           sheet[coord] = line

# save workbook
wb.save(excelName)
