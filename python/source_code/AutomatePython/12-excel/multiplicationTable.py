import openpyxl
import sys
from openpyxl.utils.cell import get_column_letter

if len(sys.argv) < 2:
    print('usage: py multiplicationTable.py [NUM]')
    exit()
num = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

for i in range(1, num+1):
    # fill the first row
    sheet[get_column_letter(i+1)+'1'] = i
    # fill the first column
    sheet['A'+str(i+1)] = i
   
# fill the table
for i in range(1, num+1):
    for j in range(1, num+1):
        sheet[get_column_letter(j+1)+str(i+1)] = i*j

wb.save('multiplicationTable-'+str(num)+'.xlsx')
