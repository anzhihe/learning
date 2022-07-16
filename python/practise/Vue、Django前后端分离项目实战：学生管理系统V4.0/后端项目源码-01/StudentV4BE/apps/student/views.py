# 引入Student的类
from student.models import Student
# 引入JsonResponse模块
from django.http import JsonResponse
# 导入json模块
import json
# 导入Q查询
from django.db.models import Q
# 导入uuid类
import uuid
# 导入哈希库
import hashlib
# 导入Setting
from django.conf import settings
# 导入os
import os
# 引入处理Excel模块
import openpyxl
# Create your views here.


def get_students(request):
    """获取所有学生的信息"""
    try:
        # 使用ORM获取所有学生信息 并把对象转为字典格式
        obj_students = Student.objects.all().values()
        # 把外层的容器转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code':1, 'data':students})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "获取学生信息出现异常，具体错误：" + str(e)})


def query_students(request):
    """查询学生信息"""
    # 接收传递过来的查询条件--- axios默认是json --- 字典类型（'inputstr'）-- data['inputstr']
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 使用ORM获取满足条件的学生信息 并把对象转为字典格式
        obj_students = Student.objects.filter(Q(sno__icontains=data['inputstr']) | Q(name__icontains=data['inputstr']) |
                                              Q(gender__icontains=data['inputstr']) | Q(mobile__icontains=data['inputstr'])
                                              | Q(email__icontains=data['inputstr']) | Q(address__icontains=data['inputstr'])).values()
        # 把外层的容器转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "查询学生信息出现异常，具体错误：" + str(e)})


def is_exists_sno(request):
    """判断学号是否存在"""
    # 接收传递过来的学号
    data = json.loads(request.body.decode('utf-8'))
    # 进行校验
    try:
        obj_students = Student.objects.filter(sno=data['sno'])
        if obj_students.count() == 0:
            return JsonResponse({'code': 1, 'exists': False})
        else:
            return JsonResponse({'code': 1, 'exists': True})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg':"校验学号失败，具体原因：" + str(e)})


def add_student(request):
    """添加学生到数据库"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 添加到数据库
        obj_student = Student(sno=data['sno'],name=data['name'],gender=data['gender'],
                              birthday=data['birthday'],mobile=data['mobile'],
                              email= data['email'], address=data['address'],image=data['image'])
        # 执行添加
        obj_student.save()
        # 使用ORM获取所有学生信息 并把对象转为字典格式
        obj_students = Student.objects.all().values()
        # 把外层的容器转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code':0 , 'msg': "添加到数据库出现异常，具体原因：" + str(e)})


def update_student(request):
    """修改学生到数据库"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 查找到要修改的学生信息
        obj_student = Student.objects.get(sno=data['sno'])
        # 依次修改
        obj_student.name = data['name']
        obj_student.gender = data['gender']
        obj_student.birthday = data['birthday']
        obj_student.mobile = data['mobile']
        obj_student.email = data['email']
        obj_student.address = data['address']
        obj_student.image = data['image']
        # 保存
        obj_student.save()
        # 使用ORM获取所有学生信息 并把对象转为字典格式
        obj_students = Student.objects.all().values()
        # 把外层的容器转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code':0 , 'msg': "修改保存到数据库出现异常，具体原因：" + str(e)})


def delete_student(request):
    """删除一条学生信息"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 查找到要修改的学生信息
        obj_student = Student.objects.get(sno=data['sno'])
        # 删除
        obj_student.delete()
        # 使用ORM获取所有学生信息 并把对象转为字典格式
        obj_students = Student.objects.all().values()
        # 把外层的容器转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "删除学生信息写入数据库出现异常，具体原因：" + str(e)})


def delete_students(request):
    """批量删除学生信息"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 遍历传递的集合
        for one_student in data['student']:
            # 查询当前记录
            obj_student = Student.objects.get(sno=one_student['sno'])
            # 执行删除
            obj_student.delete()
        # 使用ORM获取所有学生信息 并把对象转为字典格式
        obj_students = Student.objects.all().values()
        # 把外层的容器转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "批量删除学生信息写入数据库出现异常，具体原因：" + str(e)})


def upload(request):
    """接收上传的文件"""
    # 接收上传的文件
    rev_file = request.FILES.get('avatar')
    # 判断，是否有文件
    if not rev_file:
        return JsonResponse({'code':0, 'msg':'图片不存在！'})
    # 获得一个唯一的名字： uuid +hash
    new_name = get_random_str()
    # 准备写入的URL
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1] )
    # 开始写入到本次磁盘
    try:
        f = open(file_path,'wb')
        # 多次写入
        for i in rev_file.chunks():
            f.write(i)
        # 要关闭
        f.close()
        # 返回
        return JsonResponse({'code': 1, 'name': new_name + os.path.splitext(rev_file.name)[1]})

    except Exception as e:
        return JsonResponse({'code':0, 'msg':str(e)})


def get_random_str():
    #获取uuid的随机数
    uuid_val = uuid.uuid4()
    #获取uuid的随机数字符串
    uuid_str = str(uuid_val).encode('utf-8')
    #获取md5实例
    md5 = hashlib.md5()
    #拿取uuid的md5摘要
    md5.update(uuid_str)
    #返回固定长度的字符串
    return md5.hexdigest()


def import_students_excel(request):
    """从Excel批量导入学生信息"""
    # ========1. 接收Excel文件存储到Media文件夹 =======
    rev_file = request.FILES.get('excel')
    # 判断，是否有文件
    if not rev_file:
        return JsonResponse({'code': 0, 'msg': 'Excel文件不存在！'})
    # 获得一个唯一的名字： uuid +hash
    new_name = get_random_str()
    # 准备写入的URL
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])
    # 开始写入到本次磁盘
    try:
        f = open(file_path, 'wb')
        # 多次写入
        for i in rev_file.chunks():
            f.write(i)
        # 要关闭
        f.close()
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})

    #====== 2. 读取存储在Media文件夹的数据  =====
    ex_students = read_excel_dict(file_path)

    # ====3. 把读取的数据存储到数据库 =====
    # 定义几个变量： success:  error: errors
    success = 0
    error = 0
    error_snos = []

    # 开始遍历
    for one_student in ex_students:
        try:

            obj_student = Student.objects.create(sno=one_student['sno'], name=one_student['name'], gender=one_student['gender'],
                                                      birthday=one_student['birthday'],mobile=one_student['mobile'],
                                                   email=one_student['email'],address=one_student['address'])
            # 计数
            success += 1
        except:
            # 如果失败了
            error += 1
            error_snos.append(one_student['sno'])


    # 4. 返回--导入信息（成功：5，失败：4--（sno））,所有学生
    obj_students = Student.objects.all().values()
    students = list(obj_students)
    return JsonResponse({'code':1, 'success':success,'error':error,'errors':error_snos, 'data':students})

def export_student_excel(request):
    """到处数据到excel"""
    # 获取所有的学生信息
    obj_students = Student.objects.all().values()
    # 转为List
    students = list(obj_students)
    # 准备名称
    excel_name = get_random_str() + ".xlsx"
    # 准备写入的路劲
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    # 写入到Excel
    write_to_excel(students, path)
    # 返回
    return JsonResponse({'code':1, 'name':excel_name })

def read_excel_dict(path:str):
    """读取Excel数据，存储为字典 --- [{},{},{},]"""
    # 实例化一个wrokbook
    workbook = openpyxl.load_workbook(path)
    # 实例化一个sheet
    sheet = workbook['student']
    # 定义一个变量存储最终的数据--[]
    students = []
    # 准备key
    keys = ['sno','name','gender','birthday','mobile','email','address']
    # 遍历
    for row in sheet.rows:
        # 定义一个临时的字典
        temp_dict = {}
        # 组合值和key
        for index,cell in enumerate(row):
            # 组和
            temp_dict[keys[index]] = cell.value
        # 附加到list中
        students.append(temp_dict)
    #返回
    return students

def write_to_excel(data:list, path:str):
    """把数据库写入到Excel"""
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = 'student'
    # 准备keys
    keys = data[0].keys()
    # 准备写入数据
    for index, item in enumerate(data):
        # 遍历每一个元素
        for k,v in enumerate(keys):
            sheet.cell(row=index + 1, column=k+ 1, value=str(item[v]))
    # 写入到文件
    workbook.save(path)


