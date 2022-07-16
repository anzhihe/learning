import openpyxl


def read_excel_dict(path:str):
    """读取Excel数据，存储为字典 --- [{},{},{},]"""
    # 实例化一个wrokbook
    workbook = openpyxl.load_workbook(path)
    # 实例化一个sheet
    sheet = workbook['student']
    # 定义一个变量存储最终的数据--[]
    students = []
    # 准备key
    keys = ['sno','name','gender','birthday','mobile','email','address']
    # 遍历
    for row in sheet.rows:
        # 定义一个临时的字典
        temp_dict = {}
        # 组合值和key
        for index,cell in enumerate(row):
            # 组和
            temp_dict[keys[index]] = cell.value
        # 附加到list中
        students.append(temp_dict)
    #返回
    return students


def write_to_excel(data:list, path:str):
    """把数据库写入到Excel"""
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = 'student'
    # 准备keys
    keys = data[0].keys()
    # 准备写入数据
    for index in range(0, len(data)):
        # 遍历每一个元素
        for cell in range(0, len(data[index])):
            sheet.cell(row=index + 1, column=cell + 1, value=str(data[index][keys[cell]]))
    # 写入到文件
    workbook.save(path)



if __name__ == '__main__':
    path="D:\student01.xlsx"
    # 调用函数！
    students = read_excel_dict(path)
    # 输出
    print(students)